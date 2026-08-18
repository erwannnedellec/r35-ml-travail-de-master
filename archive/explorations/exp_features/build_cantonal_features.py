"""
CHANTIER FEATURES — Phase A
Construction de 8 features binaires (grain journalier) vacances scolaires et
jours feries cantonaux pour 4 cantons romands : Valais, Fribourg, Neuchatel, Geneve.

HORS CANONIQUE : ce script vit dans exp_features/ et n'altere ni le hash
416bb90b ni le ml_dataset canonique. Sortie = features candidates additionnelles,
nommees event_is_vacances_{canton} / event_is_ferie_{canton}, pretes a joindre
par date. Ne touche pas a cal_type_jour ni aux features vaudoises existantes.

Sources (data/external/) :
- Valais  : 5 PDF "Plan de scolarite Valais romand" (calendriers colories) ;
            parses via exp_features/parse_valais.py (rects vectoriels colores :
            bleu/jaune=jour de classe, orange=ferie cantonal, blanc=conge).
- Fribourg: 2 XLSX (calendrier majoritaire 20-25 + general 25-30), periodes
            explicites (partie francophone, calendrier majoritaire).
- Neuchatel: arrete 410.562 (2019-2020 a 2029-2030), table de periodes explicites.
- Geneve  : PDF "Calendrier annee scolaire" (chaque PDF liste aussi l'annee n+1).
            2024-2025 obtenu via la colonne droite du PDF 2023-2024 ; rentree 2024
            confirmee 26.08.2024 (calendrier officiel ge.ch). 2025-2026 confirme
            via PDF local (rendu image) + ge.ch.

Plage couverte : 2021-12-01 -> 2025-12-31 (dataset canonique : 2021-12-12 -> 2025-12-13).
"""
from __future__ import annotations
from datetime import date, timedelta
from pathlib import Path
import importlib.util
import pandas as pd
import pdfplumber

ROOT = Path(__file__).resolve().parents[1]
EXT = ROOT / "data" / "external"
OUT = Path(__file__).resolve().parent / "outputs"
OUT.mkdir(exist_ok=True)

CANTONS = ["valais", "fribourg", "neuchatel", "geneve"]
RANGE_START = date(2021, 12, 1)
RANGE_END = date(2025, 12, 31)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def daterange(d0: date, d1: date):
    d = d0
    while d <= d1:
        yield d
        d += timedelta(days=1)


def expand(d0: date, d1: date) -> set:
    """Toutes les dates de d0 a d1 inclus."""
    return set(daterange(d0, d1))


def easter_sunday(year: int) -> date:
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4; e = b % 4; f = (b + 8) // 25; g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30; i = c // 4; k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7; m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def federal_fast_monday(year: int) -> date:
    """Lundi du Jeune federal = lundi suivant le 3e dimanche de septembre."""
    d = date(year, 9, 1)
    first_sun = d + timedelta(days=(6 - d.weekday()) % 7)
    return first_sun + timedelta(days=14) + timedelta(days=1)


def jeune_genevois(year: int) -> date:
    """Jeudi suivant le 1er dimanche de septembre."""
    d = date(year, 9, 1)
    first_sun = d + timedelta(days=(6 - d.weekday()) % 7)
    return first_sun + timedelta(days=4)


# ---------------------------------------------------------------------------
# 1) VALAIS — parse PDF colories
# ---------------------------------------------------------------------------
def load_valais():
    spec = importlib.util.spec_from_file_location("pv", Path(__file__).parent / "parse_valais.py")
    pv = importlib.util.module_from_spec(spec); spec.loader.exec_module(pv)
    vdir = EXT / "vacances_scolaires_valais"
    files = [
        ("DEF-PlanDeScolarite2021-2022.pdf", 2021, 2022),
        ("DEF-PlanDeScolarite2022-2023.pdf", 2022, 2023),
        ("DEF-PlanDeScolarite2023-2024.pdf", 2023, 2024),
        ("DEF-PlanDeScolarite2024-2025.pdf", 2024, 2025),
        ("Plan de scolarite valais romand 2025-2026.pdf", 2025, 2026),
    ]
    parsed = []
    for fn, y1, y2 in files:
        with pdfplumber.open(vdir / fn) as doc:
            df = pv.parse_page(doc.pages[0], y1, y2)
        assert (df.row_label != df.true_wd).sum() == 0, f"weekday mismatch in {fn}"
        df["file"] = fn
        parsed.append(df)
    allp = pd.concat(parsed, ignore_index=True)
    allp["date"] = pd.to_datetime(allp["date"]).dt.date
    return allp, parsed


def valais_features():
    allp, parsed = load_valais()
    school = set(allp.loc[allp.color.isin(["blue", "yellow"]), "date"])
    orange = set(allp.loc[allp.color == "orange", "date"])

    # --- vacances : conge weekday (blanc OU orange) dans le terme scolaire,
    #     regroupe en blocs ; bloc = vacances si >= 4 jours ouvrables non-classe.
    vac = set()
    terms = []  # (first_school, last_school) par fichier
    for df in parsed:
        sch = sorted(df.loc[df.color.isin(["blue", "yellow"]), "date"].map(
            lambda d: d if isinstance(d, date) else pd.Timestamp(d).date()))
        terms.append((min(sch), max(sch)))
    school_d = {d if isinstance(d, date) else pd.Timestamp(d).date() for d in school}

    # blocs intra-terme
    for (t0, t1) in terms:
        congs = [d for d in daterange(t0, t1)
                 if d.weekday() < 5 and d not in school_d]
        # cluster consecutifs (gap <= 4 jours calendaires pour ponter les week-ends)
        clusters = []
        for d in congs:
            if clusters and (d - clusters[-1][-1]).days <= 4:
                clusters[-1].append(d)
            else:
                clusters.append([d])
        for cl in clusters:
            if len(cl) >= 4:  # vraie periode de vacances (vs pont d'1-2 jours)
                for dd in daterange(cl[0], cl[-1]):
                    vac.add(dd)

    # --- ete : entre fichiers consecutifs (couvre fin juin + juillet + debut aout)
    terms_sorted = sorted(terms)
    for (t0a, t1a), (t0b, t1b) in zip(terms_sorted, terms_sorted[1:]):
        for dd in daterange(t1a + timedelta(days=1), t0b - timedelta(days=1)):
            vac.add(dd)

    # --- feries Valais (analytique, recoupe avec orange du document) ---
    fer = set()
    for y in range(2021, 2027):
        E = easter_sunday(y)
        cand = [date(y, 1, 1), date(y, 3, 19),            # Nouvel An, St-Joseph
                E - timedelta(days=2), E + timedelta(days=1),  # Vendredi Saint, Lundi Paques
                E + timedelta(days=39), E + timedelta(days=50),  # Ascension, Lundi Pentecote
                E + timedelta(days=60),                   # Fete-Dieu
                date(y, 8, 1), date(y, 8, 15),            # Fete nationale, Assomption
                date(y, 11, 1), date(y, 12, 8), date(y, 12, 25)]  # Toussaint, Immaculee, Noel
        fer.update(cand)
    # recoupement : chaque ferie analytique en JOUR OUVRABLE present dans le doc
    # doit etre orange (controle qualite, non bloquant -> on rapporte)
    orange_d = {d if isinstance(d, date) else pd.Timestamp(d).date() for d in orange}
    doc_min = min(min(t) for t in terms); doc_max = max(max(t) for t in terms)
    check_miss = [d for d in sorted(fer) if doc_min <= d <= doc_max
                  and d.weekday() < 5 and d not in orange_d]
    return vac, fer, {"orange_confirm_misses": check_miss, "n_orange": len(orange_d)}


# ---------------------------------------------------------------------------
# 2) FRIBOURG — XLSX periodes explicites (calendrier majoritaire francophone)
# ---------------------------------------------------------------------------
def fribourg_features():
    import openpyxl
    fdir = EXT / "vacances_scolaires_fribourg"
    sheets = []  # (annee_debut, dict periodes)
    for f, yr0 in [("calendrier_20-25_majoritaire_corrigé_2.xlsx", 2020),
                   ("calendrier-general-20252030.xlsx", 2025)]:
        wb = openpyxl.load_workbook(fdir / f, data_only=True)
        for i, sh in enumerate(wb.sheetnames):
            ws = wb[sh]
            rec = {}
            for r in range(15, 27):
                b = ws.cell(r, 2).value; lab = ws.cell(r, 6).value
                if b is None or lab is None:
                    continue
                rec[str(lab).strip()] = str(b).strip()
            sheets.append((yr0 + i, rec))
    sheets = {y: r for y, r in sheets}

    def parse_dm(s, year):
        d, m = s.split(".")
        return date(year, int(m), int(d))

    def period(rec, key, ystart):
        """'DD.MM-DD.MM' -> (date_debut, date_fin) en gerant le passage d'annee."""
        s = rec[key]
        a, b = s.split("-")
        da, ma = a.split("."); db, mb = b.split(".")
        ma, mb = int(ma), int(mb)
        ya = ystart if ma >= 8 else ystart + 1
        yb = ystart if mb >= 8 else ystart + 1
        return date(ya, ma, int(da)), date(yb, mb, int(db))

    vac = set()
    debuts = {}  # year_start -> debut annee scolaire (pour calcul ete)
    derniers = {}
    for ystart, rec in sheets.items():
        # debut / dernier jour de classe
        dd = rec["Début de l'année scolaire"]; d, m = dd.split(".")
        debut = date(ystart, int(m), int(d)); debuts[ystart] = debut
        ld = rec["Dernier jour de classe"]; d, m = ld.split(".")
        dernier = date(ystart + 1, int(m), int(d)); derniers[ystart] = dernier
        for key in ["Vacances d'automne", "Vacances de Noël", "Vacances de carnaval",
                    "Vacances de Pâques"]:
            p0, p1 = period(rec, key, ystart)
            vac |= expand(p0, p1)
    # ete : dernier jour de classe (+1) -> debut annee suivante (-1)
    for ystart in sorted(derniers):
        nxt = debuts.get(ystart + 1)
        if nxt is None:
            continue
        vac |= expand(derniers[ystart] + timedelta(days=1), nxt - timedelta(days=1))

    # feries Fribourg (catholique)
    fer = set()
    for y in range(2021, 2027):
        E = easter_sunday(y)
        fer.update([date(y, 1, 1), E - timedelta(days=2), E + timedelta(days=1),
                    E + timedelta(days=39), E + timedelta(days=50), E + timedelta(days=60),
                    date(y, 8, 1), date(y, 8, 15), date(y, 11, 1),
                    date(y, 12, 8), date(y, 12, 25)])
    return vac, fer, {}


# ---------------------------------------------------------------------------
# 3) NEUCHATEL — arrete 410.562 (periodes transcrites de la table officielle)
# ---------------------------------------------------------------------------
def neuchatel_features():
    # (automne, noel, fevrier(1er mars), printemps(paques), ete) inclusifs
    P = {
        2021: [("2021-10-04", "2021-10-15"), ("2021-12-24", "2022-01-07"),
               ("2022-02-28", "2022-03-04"), ("2022-04-11", "2022-04-22"),
               ("2022-07-04", "2022-08-12")],
        2022: [("2022-10-03", "2022-10-14"), ("2022-12-26", "2023-01-06"),
               ("2023-02-27", "2023-03-03"), ("2023-04-07", "2023-04-21"),
               ("2023-07-03", "2023-08-11")],
        2023: [("2023-10-02", "2023-10-13"), ("2023-12-21", "2024-01-05"),
               ("2024-02-26", "2024-03-01"), ("2024-03-29", "2024-04-12"),
               ("2024-07-08", "2024-08-16")],
        2024: [("2024-10-07", "2024-10-18"), ("2024-12-23", "2025-01-03"),
               ("2025-02-24", "2025-02-28"), ("2025-04-14", "2025-04-25"),
               ("2025-07-07", "2025-08-15")],
        2025: [("2025-10-06", "2025-10-17"), ("2025-12-22", "2026-01-02"),
               ("2026-02-23", "2026-02-27"), ("2026-04-03", "2026-04-17"),
               ("2026-07-06", "2026-08-14")],
    }
    vac = set()
    for periods in P.values():
        for a, b in periods:
            vac |= expand(date.fromisoformat(a), date.fromisoformat(b))

    # feries Neuchatel (protestant) : 1/1, 2/1, 1er Mars, Vendredi Saint, Lundi
    # Paques, 1er Mai, Ascension, Lundi Pentecote, 1er Aout, Lundi Jeune federal, Noel
    fer = set()
    for y in range(2021, 2027):
        E = easter_sunday(y)
        fer.update([date(y, 1, 1), date(y, 1, 2), date(y, 3, 1),
                    E - timedelta(days=2), E + timedelta(days=1), date(y, 5, 1),
                    E + timedelta(days=39), E + timedelta(days=50),
                    date(y, 8, 1), federal_fast_monday(y), date(y, 12, 25)])
    return vac, fer, {}


# ---------------------------------------------------------------------------
# 4) GENEVE — PDF "Calendrier annee scolaire" (periodes transcrites)
# ---------------------------------------------------------------------------
def geneve_features():
    # (automne, noel, fevrier, paques, ete_debut, rentree_suivante) ; ete = [debut..rentree-1]
    P = {
        2021: dict(automne=("2021-10-25", "2021-10-29"), noel=("2021-12-24", "2022-01-07"),
                   fev=("2022-02-14", "2022-02-18"), paques=("2022-04-14", "2022-04-22"),
                   ete_debut="2022-07-04", rentree_next="2022-08-22"),
        2022: dict(automne=("2022-10-24", "2022-10-28"), noel=("2022-12-26", "2023-01-06"),
                   fev=("2023-02-20", "2023-02-24"), paques=("2023-04-07", "2023-04-21"),
                   ete_debut="2023-07-03", rentree_next="2023-08-21"),
        2023: dict(automne=("2023-10-23", "2023-10-27"), noel=("2023-12-25", "2024-01-05"),
                   fev=("2024-02-19", "2024-02-23"), paques=("2024-03-29", "2024-04-12"),
                   ete_debut="2024-07-01", rentree_next="2024-08-26"),
        2024: dict(automne=("2024-10-21", "2024-10-25"), noel=("2024-12-23", "2025-01-03"),
                   fev=("2025-02-24", "2025-02-28"), paques=("2025-04-18", "2025-05-02"),
                   ete_debut="2025-06-30", rentree_next="2025-08-18"),
        2025: dict(automne=("2025-10-20", "2025-10-24"), noel=("2025-12-22", "2026-01-02"),
                   fev=("2026-02-23", "2026-02-27"), paques=("2026-04-03", "2026-04-17"),
                   ete_debut="2026-06-29", rentree_next="2026-08-17"),
    }
    vac = set()
    for ystart, rec in P.items():
        for k in ["automne", "noel", "fev", "paques"]:
            a, b = rec[k]
            vac |= expand(date.fromisoformat(a), date.fromisoformat(b))
        vac |= expand(date.fromisoformat(rec["ete_debut"]),
                      date.fromisoformat(rec["rentree_next"]) - timedelta(days=1))

    # feries Geneve : 1/1, Vendredi Saint, Lundi Paques, 1er Mai, Ascension,
    # Lundi Pentecote, 1er Aout, Jeune genevois, Noel 25/12, Restauration 31/12
    fer = set()
    for y in range(2021, 2027):
        E = easter_sunday(y)
        fer.update([date(y, 1, 1), E - timedelta(days=2), E + timedelta(days=1),
                    date(y, 5, 1), E + timedelta(days=39), E + timedelta(days=50),
                    date(y, 8, 1), jeune_genevois(y), date(y, 12, 25), date(y, 12, 31)])
    return vac, fer, {}


# ---------------------------------------------------------------------------
# Assemblage
# ---------------------------------------------------------------------------
def main():
    builders = {"valais": valais_features, "fribourg": fribourg_features,
                "neuchatel": neuchatel_features, "geneve": geneve_features}
    results = {}
    for c, fn in builders.items():
        vac, fer, diag = fn()
        results[c] = (vac, fer, diag)
        print(f"[{c}] vacances={len(vac)} dates, feries={len(fer)} dates, diag={diag}")

    dim = pd.DataFrame({"date": list(daterange(RANGE_START, RANGE_END))})
    for c in CANTONS:
        vac, fer, _ = results[c]
        dim[f"event_is_vacances_{c}"] = dim["date"].isin(vac).astype("int8")
        dim[f"event_is_ferie_{c}"] = dim["date"].isin(fer).astype("int8")
    dim["date"] = pd.to_datetime(dim["date"])

    cols = ["date"] + [f"event_is_vacances_{c}" for c in CANTONS] + \
           [f"event_is_ferie_{c}" for c in CANTONS]
    dim = dim[cols]
    dim.to_parquet(OUT / "features_vacances_feries_cantonales.parquet", index=False)
    dim.to_csv(OUT / "features_vacances_feries_cantonales.csv", index=False, encoding="utf-8-sig")
    print(f"\nEcrit {len(dim)} lignes x {len(cols)} cols -> {OUT}/features_vacances_feries_cantonales.parquet")
    return dim, results


if __name__ == "__main__":
    main()
